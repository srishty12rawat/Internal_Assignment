# import openpyxl module 
import openpyxl

# Give the location of the file 
path = "death-data.xlsx"

# To open the workbook  
# workbook object is created 
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb_obj.active

# Cell object is created by using  
# sheet object's cell() method. 
cell_obj = sheet_obj.cell(row = 7, column = 8) 

# Print value of cell object  
# using the value attribute 
print(cell_obj.value) 

# print the total number of rows 
# print(sheet_obj.max_row)

# print total number of column  
# print(sheet_obj.max_column)

# max_col = sheet_obj.max_column 
  
# # Loop will print all columns name 
# for i in range(8, max_col + 1): 
#     cell_obj = sheet_obj.cell(row = 7, column = i) 
#     print(cell_obj.value) 

# m_row = sheet_obj.max_row 
  
# # Loop will print all values 
# # of first column  
# for i in range(8, m_row + 1): 
#     cell_obj = sheet_obj.cell(row = i, column = 8) 
#     print(cell_obj.value) 