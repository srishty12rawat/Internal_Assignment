# importing openpyxl module 
import openpyxl 

# Give the location of the file 
name = "death-data.xlsx"

# workbook object is created 
wb_obj = openpyxl.load_workbook(name) 

sheet_obj = wb_obj.active 

max_col = sheet_obj.max_column 
max_row = sheet_obj.max_row

# Will print a particular row value 
countries = {}
list_of_countries = []
all_causes= []
count = 1

for i in range(3, max_col + 1): 
	cell_obj = sheet_obj.cell(row = 15, column = i)
	if cell_obj.value is not None:
		# print(cell_obj.value)
		all_causes.append(cell_obj.value)

print('causes')
print(max_row)
for i in range(18, max_row):
	cell_obj = sheet_obj.cell(row = i, column = 3)
	if cell_obj.value is not None:
		print(cell_obj.value)
		# cause_obj = sheet_obj.cell(row = i , column = 3)
	# if cause_obj.value is not None:
	# print(cause_obj.value)



for i in range(7, max_col + 1): 
	cell_obj = sheet_obj.cell(row = 7, column = i)
	cell_obj2 = sheet_obj.cell(row = 8, column =i)
	cell_obj3 = sheet_obj.cell(row = 13, column =i)
	
	if cell_obj.value:
		countries["id"] = count
		countries["name"] = cell_obj.value
		countries["code"] = cell_obj2.value
		countries["population"] = cell_obj3.value
		countries["all_causes"] = all_causes[count]
		list_of_countries.append(countries)
		print(countries)
		count = count + 1


