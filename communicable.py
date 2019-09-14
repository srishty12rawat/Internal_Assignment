import xlrd
from collections import OrderedDict
import openpyxl
import simplejson as json

std_diseases = []
childhood_cluster_diseases = []
hepatitis_diseases = []
tropical_cluster_diseases = []
intestinal_nematode_diseases = []
respiratory_infection_diseases = []
perinatal_condition_diseases = []
nutritional_deficiency_diseases = []

### Nutritional deficiencies

def nutritional_deficiency(sheet, nutritional_deficiency):
    data = OrderedDict()
    for i in range(62, 66):
        data[sheet_obj.cell(i,6).value] = sheet_obj.cell(i,8).value
    nutritional_deficiency.append(data) 
    j = json.dumps(nutritional_deficiency)       
    with open('nutritional_deficiency.json', 'w') as f:
        f.write(j) 
    return nutritional_deficiency


### Perinatal conditions

def perinatal_condition(sheet, perinatal_condition):
    data = OrderedDict()
    for i in range(58, 61):
        data[sheet_obj.cell(i,6).value] = sheet_obj.cell(i,8).value
    perinatal_condition.append(data) 
    j = json.dumps(perinatal_condition)       
    with open('perinatal_condition.json', 'w') as f:
        f.write(j) 
    return perinatal_condition


### Respiratory infection

def respiratory_infection(sheet, respiratory_infection):
    data = OrderedDict()
    for i in range(58, 61):
        data[sheet_obj.cell(i,6).value] = sheet_obj.cell(i,8).value
    respiratory_infection.append(data) 
    j = json.dumps(respiratory_infection)       
    with open('respiratory_infection.json', 'w') as f:
        f.write(j) 
    return respiratory_infection

### Intestinal nematode

def intestinal_nematode(sheet, intestinal_nematode):
    data = OrderedDict()
    for i in range(49, 52):
        data[sheet_obj.cell(i,6).value] = sheet_obj.cell(i,8).value
    intestinal_nematode.append(data) 
    j = json.dumps(intestinal_nematode)       
    with open('intestinal_nematode.json', 'w') as f:
        f.write(j) 
    return intestinal_nematode

### Tropical-cluster 

def tropical_cluster(sheet, tropical_cluster):
    data = OrderedDict()
    for i in range(38, 44):
        data[sheet_obj.cell(i,6).value] = sheet_obj.cell(i,8).value
    tropical_cluster.append(data) 
    j = json.dumps(tropical_cluster)       
    with open('tropical_cluster.json', 'w') as f:
        f.write(j) 
    return tropical_cluster

### Childhood cluster 

def childhood_cluster(sheet, childhood_cluster):
    data = OrderedDict()
    for i in range(28, 32):
        data[sheet_obj.cell(i,6).value] = sheet_obj.cell(i,8).value
    childhood_cluster.append(data) 
    j = json.dumps(childhood_cluster)       
    with open('childhood_cluster.json', 'w') as f:
        f.write(j) 
    return childhood_cluster    

### STD excluding HIV 
def std(sheet, std):
    data = OrderedDict()
    for i in range(22, 25):
        data[sheet_obj.cell(i,6).value] = sheet_obj.cell(i,8).value
    std.append(data) 
    j = json.dumps(std)       
    with open('std.json', 'w') as f:
        f.write(j) 
    return std   


#### Hepatitis
def hepatitis(sheet, hepatitis):
    data = OrderedDict()
    for i in range(34, 36):
        data[sheet_obj.cell(i,6).value] = sheet_obj.cell(i,8).value
    hepatitis.append(data) 
    j = json.dumps(hepatitis)       
    with open('hepatitis.json', 'w') as f:
        f.write(j) 
    return hepatitis     

def infectious_parasite(sheet,infectious_parasite):
    data = OrderedDict()
    std(sheet_obj, std_diseases)
    childhood_cluster(sheet_obj, childhood_cluster_diseases)
    hepatitis(sheet_obj, hepatitis_diseases)
    tropical_cluster(sheet_obj, tropical_cluster_diseases)
    intestinal_nematode(sheet_obj, intestinal_nematode_diseases)

    data[sheet_obj.cell(20,6)] = sheet_obj(20,8).value
    data[sheet_obj.cell(21,6)] = std_diseases
    data[sheet_obj.cell(25,6)] = sheet_obj(25,8).value
    data[sheet_obj.cell(26,6)] = sheet_obj(26,8).value
    data[sheet_obj.cell(27,6)] = childhood_cluster_diseases
    data[sheet_obj.cell(33,6)] = sheet_obj(33,8).value
    data[sheet_obj.cell(34,6)] = hepatitis
    data[sheet_obj.cell(36,6)] = sheet_obj(36,8).value
    data[sheet_obj.cell(37,6)] = tropical_cluster_diseases
    data[sheet_obj.cell(44,6)] = sheet_obj(44,8).value
    data[sheet_obj.cell(45,6)] = sheet_obj(45,8).value
    data[sheet_obj.cell(46,6)] = sheet_obj(46,8).value
    data[sheet_obj.cell(47,6)] = sheet_obj(47,8).value
    data[sheet_obj.cell(48,6)] = intestinal_nematode_diseases

    infectious_parasite.append(data) 
    j = json.dumps(infectious_parasite)       
    with open('infectious_parasite.json', 'w') as f:
        f.write(j) 
    return infectious_parasite
   