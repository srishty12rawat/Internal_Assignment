import xlrd
from collections import OrderedDict
import openpyxl
import simplejson as json

### This is the intentional part

unintentional_injuries = []
intentional_injuries = []
Injuries = []


def intentional(sheet, int_inj):
        # Iterate through each row in worksheet and fetch values into dict
        # for rownum in range(1, sh.nrows):
        data = OrderedDict()
            # row_values = sh.row_values(rownum)
        data['Self-inflicted_injuries'] = sheet_obj.cell(142,8).value
        data['Violence'] = sheet_obj.cell(143,8).value
        data['War'] = sheet_obj.cell(144,8).value
        return intentional_injuries.append(data)
        
        # intentional_injuries = {'intentional_injuries': intentional_injuries}  

        # Serialize the list of dicts to JSON
        j = json.dumps(intentional_injuries)
        # Write to file
        with open('intentional_injuries.json', 'w') as f:
            f.write(j)  

#### This is the unintentional part

def unintentional(sheet, unint_inj):


        data = OrderedDict()
        data['Road_traffic_accidents'] = sheet_obj.cell(135,8).value
        data['Poisonings'] = sheet_obj.cell(136,8).value
        data['Falls'] = sheet_obj.cell(137,8).value
        data['Fires'] = sheet_obj.cell(138,8).value
        data['Drownings'] = sheet_obj.cell(139,8).value
        data['Others'] = sheet_obj.cell(140,8).value
        return unintentional_injuries.append(data)


        # Serialize the list of dicts to JSON
        j = json.dumps(unintentional_injuries)
        # Write to file
        with open('Unintentional_injuries.json', 'w') as f:
            f.write(j)


######This is the final injuries part
def injuries(sheet):

    global unintentional_injuries
    global intentional_injuries
    global Injuries


    intentional(sheet, intentional_injuries)
    unintentional(sheet, unintentional_injuries)

    data = OrderedDict()

    data['Unintentional'] = unintentional_injuries
    data['Intentional'] = intentional_injuries
    Injuries.append(data)

    Injuries = {'Injuries': Injuries}  

    # Serialize the list of dicts to JSON
    j = json.dumps(Injuries)
    # Write to file
    with open('Injuries.json', 'w') as f:
        f.write(j)  



if __name__ == "__main__":
    path = "death-data.xlsx"

    # To open the workbook  
    # workbook object is created 
    wb_obj = openpyxl.load_workbook(path)

    # Get workbook active sheet object 
    # from the active attribute 
    sheet_obj = wb_obj.active
    injuries(sheet_obj)


    
