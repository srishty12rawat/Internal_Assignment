import xlrd
from collections import OrderedDict
import openpyxl
import simplejson as json

malignant_diseases = []
neuropsychiatric_diseases = []
sense_organ_diseases = []
cardiovascular_diseases = []
respiratory_diseases = []
digestive_diseases = []
genitourinary_diseases = []
musculoskeletal_diseases = []
oral_diseases = []
nonCommunicable = []

### Mailgnant diseases

def malignant(sheet, malignant):
    data = OrderedDict()
    for i in range(68, 84):
        data[sheet_obj.cell(i,6).value] = sheet_obj.cell(i,8).value
    malignant.append(data) 
    j = json.dumps(malignant)       
    with open('malignant.json', 'w') as f:
        f.write(j) 
    return malignant

### neuro diseases
def neuropsychiatric(sheet, neuropsychiatric):
    data = OrderedDict()
    for i in range(88, 102):
        data[sheet_obj.cell(i,6).value] = sheet_obj.cell(i,8).value
    neuropsychiatric.append(data) 
    j = json.dumps(neuropsychiatric)       
    with open('neuropsychiatric.json', 'w') as f:
        f.write(j)
    return neuropsychiatric  


##### sense organ 
def sense_organ(sheet, sense_organ):
    data = OrderedDict()
    for i in range(103, 108):
        data[sheet_obj.cell(i,6).value] = sheet_obj.cell(i,8).value
    sense_organ.append(data) 
    j = json.dumps(sense_organ)       
    with open('sense_organ.json', 'w') as f:
        f.write(j)
    return sense_organ        

##### cardiovascular 
def cardiovascular(sheet, cardiovascular):
    data = OrderedDict()
    for i in range(109, 114):
        data[sheet_obj.cell(i,6).value] = sheet_obj.cell(i,8).value
    cardiovascular.append(data) 
    j = json.dumps(cardiovascular)       
    with open('cardiovascular.json', 'w') as f:
        f.write(j)
    return cardiovascular

#### respiratory
def respiratory(sheet, respiratory):
    data = OrderedDict()
    for i in range(109, 114):
        data[sheet_obj.cell(i,6).value] = sheet_obj.cell(i,8).value
    respiratory.append(data) 
    j = json.dumps(respiratory)       
    with open('respiratory.json', 'w') as f:
        f.write(j)
    return respiratory

#### digestive
def digestive(sheet, digestive):
    data = OrderedDict()
    for i in range(118, 121):
        data[sheet_obj.cell(i,6).value] = sheet_obj.cell(i,8).value
    digestive.append(data) 
    j = json.dumps(digestive)       
    with open('digestive.json', 'w') as f:
        f.write(j)
    return digestive

#### genitourinary
def genitourinary(sheet, genitourinary):
    data = OrderedDict()
    for i in range(122, 124):
        data[sheet_obj.cell(i,6).value] = sheet_obj.cell(i,8).value
    genitourinary.append(data) 
    j = json.dumps(genitourinary)       
    with open('genitourinary.json', 'w') as f:
        f.write(j)
    return genitourinary  

#### musculoskeletal
def musculoskeletal(sheet, musculoskeletal):
    data = OrderedDict()
    for i in range(126, 128):
        data[sheet_obj.cell(i,6).value] = sheet_obj.cell(i,8).value
    musculoskeletal.append(data) 
    j = json.dumps(musculoskeletal)       
    with open('musculoskeletal.json', 'w') as f:
        f.write(j)
    return musculoskeletal 


#### oral
def oral(sheet, oral):
    data = OrderedDict()
    for i in range(130, 133):
        data[sheet_obj.cell(i,6).value] = sheet_obj.cell(i,8).value
    oral.append(data) 
    j = json.dumps(oral)       
    with open('oral.json', 'w') as f:
        f.write(j)
    return oral     


### Non-communicable
def noncom(sheet):
    global malignant_diseases
    global neuropsychiatric_diseases
    global sense_organ_diseases
    global cardiovascular_diseases
    global respiratory_diseases
    global digestive_diseases
    global genitourinary_diseases
    global musculoskeletal_diseases
    global oral_diseases
    global nonCommunicable

    malignant(sheet_obj,malignant_diseases)
    neuropsychiatric(sheet_obj,neuropsychiatric_diseases)
    sense_organ(sheet_obj,sense_organ_diseases)
    cardiovascular(sheet_obj,cardiovascular_diseases)
    respiratory(sheet_obj,respiratory_diseases)
    digestive(sheet_obj,digestive_diseases)
    genitourinary(sheet_obj,genitourinary_diseases)
    musculoskeletal(sheet_obj,musculoskeletal_diseases)
    oral(sheet_obj,oral_diseases)

    data = OrderedDict()

    data['Malignant neoplasms'] = malignant_diseases
    data['Other neoplasms'] = sheet_obj.cell(84,8).value
    data['Diabetes mellitus'] = sheet_obj.cell(85,8).value
    data['Endocrine disorders'] = sheet_obj.cell(86,8).value
    data['Neuropsychiatric conditions'] = neuropsychiatric_diseases
    data['Sense organ diseases'] = sense_organ_diseases
    data['Cardiovascular diseases'] = cardiovascular_diseases
    data['Respiratory diseases'] = respiratory_diseases
    data['Digestive diseases'] = digestive_diseases
    data['Genitourinary diseases'] = genitourinary_diseases
    data['Skin diseases'] = sheet_obj.cell(124,8).value
    data['Musculoskeletal diseases'] = musculoskeletal_diseases
    data['Congenital anomalies'] = sheet_obj.cell(128,8).value
    data['Oral conditions'] = oral_diseases

    nonCommunicable.append(data)

    nonCommunicable = {'nonCommunicable': nonCommunicable}  

    # Serialize the list of dicts to JSON
    j = json.dumps(nonCommunicable)
    # Write to file
    with open('nonCommunicable.json', 'w') as f:
        f.write(j)   

if __name__ == "__main__":
    path = "death-data.xlsx"    
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    noncom(sheet_obj)

    



